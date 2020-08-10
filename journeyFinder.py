#! /usr/bin/env python3
'''
Raymond Yu
Script uses the openpyxl and selenium libaries of python
in order to scrape data about points of interest and restaurants
from TripAdvisor to put into an excel worksheet
'''
import time, os
from openpyxl import Workbook
from selenium import webdriver

PATH = '/usr/local/bin/chromedriver'
'''Prompt Input'''
yes = False
acceptable = ['y', 'Y', 'Yes', 'yes']
print('Outputted Excel files will be in ./files')
while not yes:
    print('What location would you like to visit?')
    print('Format: City State/Country') 
    print('Example: Dallas Texas, London United Kingdom')
    location = input('Enter: ')
    confirm = input('Are you sure? (y/n) ')
    if confirm in acceptable:
        yes = True

'''Setup driver for Chrome'''
driver = webdriver.Chrome(PATH) 
driver.set_window_size(1200, 700)
driver.implicitly_wait(10)
driver.get('http://www.tripadvisor.com/') 

'''Search City'''
elements = driver.find_elements_by_name('q')
search = elements[1]
search.click()
search.send_keys(location)
search.submit()

'''Overview of City'''
results = driver.find_elements_by_class_name('result-title')
city = results[0]
city.click()

'''Change Window'''
allWindows = driver.window_handles
parentWindow = allWindows[0]
currWindow = allWindows[1]
driver.switch_to.window(currWindow)

'''Go to City'''
url = driver.current_url
url_elem = url.split('-')
city_id = url_elem[1]
city_name = url_elem[2]
attraction_url = 'https://www.tripadvisor.com/Attractions-' + city_id + '-Activities-oa30-' + city_name + '.html'
restaurant_url = 'https://www.tripadvisor.com/Restaurants-' + city_id + '-' + city_name + '.html'

'''Collect attractions'''
names = []
categories = []
reviews = []
urls = []

driver.get(attraction_url)
i = 0
while i < 2:
    pageNums = driver.find_elements_by_class_name('pageNum')
    if len(pageNums) != 1:
        driver.execute_script('arguments[0].scrollIntoView();', pageNums[i])
        driver.execute_script('window.scrollBy(0, 90);')
        pageNums[i].click()
    else:
        previous = driver.find_elements_by_class_name('previous')
        previous[0].click()
        i = 1
    cards = driver.find_elements_by_class_name('_6sUF3jUd')
    for c in cards:
        category = c.find_elements_by_class_name('_21qUqkJx')
        if category:
            categories.append(category[0].text)
        else:
            categories.append('NOT FOUND')
        url = c.find_element_by_class_name('_1QKQOve4')
        urls.append(url.get_attribute('href'))
        place = url.find_element_by_tag_name('h2')
        names.append(place.text)
        review = c.find_elements_by_class_name('_1KK223I5')
        if review:
            reviews.append(review[0].text)
        else:
            reviews.append('NOT FOUND')
    i += 1
# print(names)
# print(categories)
# print(urls)
# print(reviews)

'''Collect Restaurants'''
driver.get(restaurant_url)

dining_names = []
dining_categories = []
dining_prices = []
dining_reviews = []
dining_urls = []

i = 1
while True:
    time.sleep(3)
    cards = driver.find_elements_by_class_name('_2Q7zqOgW')
    for c in cards:
        d = c.find_element_by_class_name('_15_ydu6b')
        dining_names.append(d.text)
        dining_urls.append(d.get_attribute('href'))
        r = c.find_elements_by_class_name('w726Ki5B')
        if r:
            dining_reviews.append(r[0].text)
        else:
            dining_reviews.append('NOT FOUND')
        #Possible optimization?
        style = c.find_element_by_class_name('_3d9EnJpt')
        stats = style.find_elements_by_class_name('EHA742uW')
        if len(stats) < 2:
            if len(stats) == 0:
                dining_categories.append('NOT FOUND')
                dining_prices.append('NOT FOUND')
            elif stats[0].text[0] == '$':
                dining_categories.append('NOT FOUND')
                dining_prices.append(stats[0].text)
            else:
                dining_categories.append(stats[0].text)
                dining_prices.append('NOT FOUND')
        else:
            dining_categories.append(stats[0].text)
            dining_prices.append(stats[1].text)
    i += 1
    if i == 5:
        break
    time.sleep(2)
    pageNums = driver.find_elements_by_class_name('pageNum')
    if i-1 >= len(pageNums):
        break
    driver.execute_script('arguments[0].scrollIntoView();', pageNums[i-1])
    driver.execute_script('window.scrollBy(0, 90);')
    pageNums[i-1].click()

# print(dining_names)
'''Quit driver'''
driver.quit()

'''Write to Excel file'''
if names:
    os.chdir('./files/')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Points of Interest'
    ws['A1'] = 'Location Name'
    ws['B1'] = 'Type'
    ws['C1'] = 'Review Count'
    ws['D1'] = 'Links'
    for x in range(len(names)):
        ws.cell(row=x+2,column=1,value=names[x])
        ws.cell(row=x+2,column=2,value=categories[x])
        ws.cell(row=x+2,column=3,value=reviews[x])
        ws.cell(row=x+2,column=4,value=urls[x])
    if dining_names:
        ws1 = wb.create_sheet('Restaurants')
        ws1['A1'] = 'Restaurant Name'
        ws1['B1'] = 'Cuisine Categories'
        ws1['C1'] = 'Price Level'
        ws1['D1'] = 'Review Count'
        ws1['E1'] = 'Links'
        for y in range(len(dining_names)):
            ws1.cell(row=y+2,column=1,value=dining_names[y])
            ws1.cell(row=y+2,column=2,value=dining_categories[y])
            ws1.cell(row=y+2,column=3,value=dining_prices[y])
            ws1.cell(row=y+2,column=4,value=dining_reviews[y])
            ws1.cell(row=y+2,column=5,value=dining_urls[y])
    excel_name = location + ' Tourism.xlsx'
    wb.save(excel_name)